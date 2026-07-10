"""
Storefront Live Activity — Redis entegreli versiyon.
Bağımsız servis olarak çalışır, mevcut shoptimize backend'e bağımlılık yok.
"""

import asyncio
import json
import logging
import os
import secrets
import time
from typing import Optional

import httpx
import requests
from fastapi import APIRouter, Depends, Query, Request
from services.auth import get_current_user
from fastapi.responses import JSONResponse, Response, StreamingResponse

SHOPIFY_API_VERSION = os.getenv("SHOPIFY_API_VERSION", "2026-04")
from services.db import get_setting, set_connection_settings
from services.redis_store import store

logger = logging.getLogger(__name__)
router = APIRouter()

API_BASE_URL = os.getenv("API_BASE_URL", "https://api.shoptimize.com.tr").rstrip("/")
# SHOPIFY_APP_URL is the public URL of this service (e.g. https://live.shoptimize.com.tr).
# We also check API_BASE_URL for legacy script tags installed when the two vars differed.
_SHOPIFY_APP_URL = os.getenv("SHOPIFY_APP_URL", "https://live.shoptimize.com.tr").rstrip("/")

# ---------------------------------------------------------------------------
# Geriye dönük uyumluluk
# ---------------------------------------------------------------------------

def register_tid_owner(tid: str, username: str, brand: str):
    if not tid:
        return
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            asyncio.ensure_future(store.register_tid_owner(tid, username, brand))
        else:
            loop.run_until_complete(store.register_tid_owner(tid, username, brand))
    except Exception as e:
        logger.warning("[LIVE] register_tid_owner sync wrapper hatası: %s", e)


# ---------------------------------------------------------------------------
# pixel.js
# ---------------------------------------------------------------------------

_BOT_UA_KEYWORDS = (
    "googlebot", "adsbot-google", "mediapartners-google",
    "bingbot", "msnbot", "slurp", "duckduckbot",
    "baiduspider", "yandexbot", "facebookexternalhit",
    "twitterbot", "linkedinbot", "whatsapp",
    "semrushbot", "ahrefsbot", "mj12bot", "dotbot",
    "rogerbot", "seznambot", "pinterestbot", "applebot",
    "python-requests", "go-http-client", "axios/",
    "curl/", "wget/", "scrapy", "lighthouse",
)


def _is_bot(ua: str) -> bool:
    ua_lower = ua.lower()
    return any(k in ua_lower for k in _BOT_UA_KEYWORDS)


_PIXEL_JS_TEMPLATE = """
/* Shoptimize Storefront Pixel v2 */
if (window._spt_loaded) { /* already loaded */ } else {
window._spt_loaded = true;
(function () {
  'use strict';

  var TID = '{TID}';
  var API = '{API}';

  if (!TID || TID === '__NOTID__') {
    console.warn('[SPT] Tracking ID bulunamadı, pixel devre dışı.');
    return;
  }

  var VID = '';
  try {
    VID = localStorage.getItem('_spt_vid') || '';
    if (!VID) {
      VID = Math.random().toString(36).slice(2, 10) + Date.now().toString(36);
      localStorage.setItem('_spt_vid', VID);
    }
  } catch (e) {
    VID = Math.random().toString(36).slice(2, 10);
  }

  var UA = navigator.userAgent || '';
  var SW = screen.width || 0;

  var UTM = {};
  try {
    var _storedUtm = sessionStorage.getItem('_spt_utm');
    if (_storedUtm) { try { UTM = JSON.parse(_storedUtm); } catch(e) {} }
    var _usp = new URLSearchParams(location.search);
    var _hasNew = false;
    ['utm_source','utm_medium','utm_campaign','utm_content','utm_term'].forEach(function(k) {
      var v = _usp.get(k); if (v) { UTM[k] = v; _hasNew = true; }
    });
    // fbclid / gclid — UTM yoksa reklam platformunu otomatik tespit et
    if (!UTM.utm_source) {
      var _fbc = _usp.get('fbclid');
      var _gc  = _usp.get('gclid');
      if (_fbc) { UTM.utm_source = 'facebook'; UTM.utm_medium = 'paid'; UTM.utm_campaign = UTM.utm_campaign || 'Facebook Ads'; _hasNew = true; }
      else if (_gc) { UTM.utm_source = 'google'; UTM.utm_medium = 'cpc'; UTM.utm_campaign = UTM.utm_campaign || 'Google Ads'; _hasNew = true; }
    }
    if (_hasNew) sessionStorage.setItem('_spt_utm', JSON.stringify(UTM));
  } catch(e) {}

  var CID = '';
  try {
    CID = String(
      (window.ShopifyAnalytics && window.ShopifyAnalytics.meta &&
       window.ShopifyAnalytics.meta.page && window.ShopifyAnalytics.meta.page.customerId) ||
      (window.__st && window.__st.cid) || ''
    );
  } catch(e) {}

  function send(event_type, data) {
    var payload = JSON.stringify({
      tid: TID, vid: VID, event_type: event_type,
      url: location.href, referrer: document.referrer || '',
      ts: Date.now(), ua: UA, sw: SW,
      utm: UTM, customer_id: CID,
      data: data || {}
    });
    try {
      fetch(API + '/api/live/event', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: payload,
        keepalive: true,
        mode: 'cors',
      }).catch(function () {});
    } catch (e) {
      try {
        var xhr = new XMLHttpRequest();
        xhr.open('POST', API + '/api/live/event', true);
        xhr.setRequestHeader('Content-Type', 'application/json');
        xhr.send(payload);
      } catch (e2) {}
    }
  }

  function getProductMeta() {
    try {
      var m = window.ShopifyAnalytics && window.ShopifyAnalytics.meta;
      if (m && m.product && m.product.title) {
        // featured_image: eski temalarda string, yeni OS2.0 temalarında {src, width, height} objesi
        var _fi = m.product.featured_image;
        var _img = typeof _fi === 'string' ? _fi : (_fi && typeof _fi === 'object' ? (_fi.src || '') : '');
        // images array'i dene (daha güvenilir)
        if (!_img && m.product.images && m.product.images.length) {
          var _i0 = m.product.images[0];
          _img = typeof _i0 === 'string' ? _i0 : (_i0 && _i0.src) || '';
        }
        return {
          product_id:     String(m.product.id || ''),
          product_title:  m.product.title,
          product_handle: m.product.handle || '',
          product_price:  m.product.price ? (m.product.price / 100).toFixed(2) : null,
          product_vendor: m.product.vendor || '',
          product_type:   m.product.type   || '',
          product_image:  _img
        };
      }
    } catch (e) {}
    try {
      var scripts = document.querySelectorAll('script[type="application/ld+json"]');
      for (var i = 0; i < scripts.length; i++) {
        var ld = JSON.parse(scripts[i].textContent);
        var item = ld['@type'] === 'Product' ? ld : (ld['@graph'] || []).find(function(x){ return x['@type'] === 'Product'; });
        if (item && item.name) {
          var imgs = item.image || [];
          var img  = typeof imgs === 'string' ? imgs : (imgs[0] || '');
          var offr = item.offers || {};
          if (Array.isArray(offr)) offr = offr[0] || {};
          return {
            product_id: '', product_title: item.name,
            product_handle: location.pathname.split('/products/')[1] || '',
            product_price: offr.price || null,
            product_vendor: (item.brand && item.brand.name) || '',
            product_type: item.category || '', product_image: img
          };
        }
      }
    } catch (e) {}
    try {
      var ogTitle = document.querySelector('meta[property="og:title"]');
      var ogImage = document.querySelector('meta[property="og:image"]');
      var ogPrice = document.querySelector('meta[property="product:price:amount"]');
      var ogBrand = document.querySelector('meta[property="og:brand"]');
      if (ogTitle && ogTitle.content) return {
        product_id: '', product_title: ogTitle.content,
        product_handle: location.pathname.split('/products/')[1] || '',
        product_price: ogPrice ? ogPrice.content : null,
        product_vendor: ogBrand ? ogBrand.content : '',
        product_type: '', product_image: ogImage ? ogImage.content : ''
      };
    } catch (e) {}
    try {
      var handle = location.pathname.split('/products/')[1];
      if (handle) {
        handle = handle.split('?')[0].split('/')[0];
        return {
          product_id: '', product_title: document.title.replace(/ [\\u2013\\-|].*/,'').trim() || handle,
          product_handle: handle, product_price: null,
          product_vendor: '', product_type: '', product_image: ''
        };
      }
    } catch (e) {}
    return {};
  }

  function fetchCartAndSend(base) {
    try {
      fetch('/cart.js', { credentials: 'same-origin' })
        .then(function(r) { return r.json(); })
        .then(function(cart) {
          var items = (cart.items || []).slice(0, 15).map(function(i) {
            return {
              title: i.title || i.product_title || '',
              price: i.price ? (i.price / 100).toFixed(2) : null,
              quantity: i.quantity || 1,
              variant_id: String(i.variant_id || ''),
              image: i.image || i.featured_image && i.featured_image.url || ''
            };
          });
          send('cart_viewed', Object.assign({}, base, {
            page_type: 'cart',
            cart_total: cart.total_price ? (cart.total_price / 100).toFixed(2) : null,
            cart_item_count: cart.item_count || 0,
            cart_items: items
          }));
        })
        .catch(function() { send('cart_viewed', Object.assign({}, base, { page_type: 'cart' })); });
    } catch(e) { send('cart_viewed', Object.assign({}, base, { page_type: 'cart' })); }
  }

  function trackProductPage(base, meta) {
    // /products/{handle}.js — Shopify storefront public API, auth gerektirmez
    // ShopifyAnalytics.meta'da görsel gelmiyorsa buradan çekiyoruz
    var handle = meta.product_handle || '';
    if (!handle || meta.product_image) {
      send('product_viewed', Object.assign({}, base, { page_type: 'product' }, meta));
      return;
    }
    fetch('/products/' + handle + '.js', { credentials: 'same-origin' })
      .then(function(r) { return r.json(); })
      .then(function(p) {
        var img = '';
        var fi = p.featured_image;
        if (typeof fi === 'string' && fi) img = fi;
        else if (fi && fi.src) img = fi.src;
        if (!img && p.images && p.images.length) {
          var i0 = p.images[0];
          img = typeof i0 === 'string' ? i0 : (i0 && i0.src) || '';
        }
        send('product_viewed', Object.assign({}, base, { page_type: 'product' }, meta, { product_image: img }));
      })
      .catch(function() {
        send('product_viewed', Object.assign({}, base, { page_type: 'product' }, meta));
      });
  }

  function trackPageView() {
    var path = location.pathname;
    var base = { title: document.title, path: path };
    if      (path.indexOf('/products/')    !== -1) trackProductPage(base, getProductMeta());
    else if (path.indexOf('/collections/') !== -1) {
      var colHandle = path.split('/collections/')[1] || '';
      colHandle = colHandle.split('/')[0].split('?')[0];
      send('collection_viewed', Object.assign({}, base, { page_type:'collection', collection_handle: colHandle }));
    }
    else if (path.match(/^\\/cart/))               fetchCartAndSend(base);
    else if (path.indexOf('/search')       !== -1) {
      var q = ''; try { q = new URLSearchParams(location.search).get('q') || ''; } catch(e){}
      send('search_submitted', Object.assign({}, base, { page_type:'search', query:q }));
    }
    else                                           send('page_viewed', Object.assign({}, base, { page_type:'page' }));
  }

  if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', trackPageView);
  } else {
    trackPageView();
  }

  function parseCartBody(body) {
    if (!body) return {};
    if (typeof body === 'string') {
      try { return JSON.parse(body); } catch(e) {}
      try { var usp = new URLSearchParams(body); return { id: usp.get('id'), quantity: usp.get('quantity') }; } catch(e) {}
    }
    if (typeof body.get === 'function') return { id: body.get('id'), quantity: body.get('quantity') };
    return {};
  }

  function cartEventData(d) {
    var items = d.items;
    var id    = d.id || (items && items[0] && items[0].id);
    var qty   = d.quantity || (items && items[0] && items[0].quantity) || 1;
    return { variant_id: id, quantity: qty };
  }

  var _origFetch = window.fetch;
  window.fetch = function(resource, init) {
    var url = String(typeof resource === 'string' ? resource : (resource && resource.url) || '');
    if (url.indexOf('/cart/add') !== -1) {
      try { send('add_to_cart', cartEventData(parseCartBody((init || {}).body))); } catch(e) {}
    }
    return _origFetch.apply(this, arguments);
  };

  var _origOpen = XMLHttpRequest.prototype.open;
  var _origSend = XMLHttpRequest.prototype.send;
  XMLHttpRequest.prototype.open = function(method, url) {
    this._spt_url = String(url || '');
    return _origOpen.apply(this, arguments);
  };
  XMLHttpRequest.prototype.send = function(body) {
    if (this._spt_url && this._spt_url.indexOf('/cart/add') !== -1) {
      try { send('add_to_cart', cartEventData(parseCartBody(body))); } catch(e) {}
    }
    return _origSend.apply(this, arguments);
  };

  document.addEventListener('submit', function(e) {
    var f = e.target;
    if (f && f.action && String(f.action).indexOf('/cart/add') !== -1) {
      try {
        var idEl = f.querySelector('[name="id"]');
        var qEl  = f.querySelector('[name="quantity"]');
        send('add_to_cart', { variant_id: idEl ? idEl.value : null, quantity: qEl ? (parseInt(qEl.value) || 1) : 1 });
      } catch(e) {}
    }
  }, true);

  document.addEventListener('click', function(e) {
    var el = e.target;
    for (var i = 0; i < 8 && el && el.tagName !== 'BODY'; i++) {
      var href = (el.getAttribute && el.getAttribute('href')) || '';
      var name = (el.getAttribute && el.getAttribute('name')) || '';
      var cls  = (el.className && typeof el.className === 'string') ? el.className : '';
      if (href.indexOf('/checkout') !== -1 || name === 'checkout' || cls.indexOf('checkout') !== -1) {
        send('checkout_started', {}); break;
      }
      el = el.parentElement;
    }
  }, true);

  document.addEventListener('cart:item-added', function(e) {
    try { send('add_to_cart', { product_title: e.detail && e.detail.product && e.detail.product.title }); } catch(er) {}
  });

  // ── Scroll depth tracking (ürün sayfalarında) ─────────────────────────────
  (function () {
    var isProduct = location.pathname.indexOf('/products/') !== -1;
    if (!isProduct) return;
    var marks = { 25: false, 50: false, 75: false, 100: false };
    var timer = null;
    function checkScroll() {
      var scrolled = window.scrollY + window.innerHeight;
      var total = Math.max(document.documentElement.scrollHeight, 1);
      var pct = Math.round(scrolled / total * 100);
      [25, 50, 75, 100].forEach(function(mark) {
        if (!marks[mark] && pct >= mark) {
          marks[mark] = true;
          send('scroll_depth', { depth: mark, page_type: 'product' });
        }
      });
    }
    window.addEventListener('scroll', function() {
      if (timer) clearTimeout(timer);
      timer = setTimeout(checkScroll, 150);
    }, { passive: true });
    // İlk yüklemede kontrol (sayfa kısa olabilir)
    setTimeout(checkScroll, 500);
  })();

  // ── Attention time (sayfada geçirilen süre) ───────────────────────────────
  (function () {
    var pageType = location.pathname.indexOf('/products/') !== -1 ? 'product'
                 : location.pathname.match(/^\\/cart/) ? 'cart'
                 : location.pathname.indexOf('/collections/') !== -1 ? 'collection'
                 : location.pathname.indexOf('/checkout') !== -1 ? 'checkout'
                 : 'page';
    var startTs = Date.now();
    var accumulated = 0;
    var sent = false;

    document.addEventListener('visibilitychange', function() {
      if (document.hidden) {
        accumulated += Date.now() - startTs;
      } else {
        startTs = Date.now();
      }
    });

    function sendAttention() {
      if (sent) return;
      sent = true;
      var total = accumulated + (document.hidden ? 0 : Date.now() - startTs);
      if (total >= 3000) { // en az 3 saniye
        send('attention_time', {
          seconds: Math.round(total / 1000),
          page_type: pageType
        });
      }
    }

    window.addEventListener('beforeunload', sendAttention);
    // Pagehide — iOS Safari için (beforeunload her zaman tetiklenmez)
    window.addEventListener('pagehide', sendAttention);
  })();

  console.info('[SPT] Shoptimize pixel aktif. TID:', TID.slice(-8));
})();
} /* end _spt_loaded guard */
"""


@router.get("/pixel.js")
async def serve_pixel(tid: str = Query(""), shop: str = Query(""), request: Request = None):
    tracking_id = tid.strip()

    # Theme App Extension path: shop= parametresi ile TID'yi DB'den çöz
    if not tracking_id and shop:
        shop_clean = shop.strip().replace("https://", "").replace("http://", "").rstrip("/")
        try:
            from services.db import lookup_username_by_shop
            found = lookup_username_by_shop(shop_clean)
            if found:
                _u, _b = found
                tracking_id = get_setting(_u, _b, "shopify", "pixel_tracking_id", "")
        except Exception:
            pass

    tracking_id = tracking_id or "__NOTID__"
    js = _PIXEL_JS_TEMPLATE.replace("{TID}", tracking_id).replace("{API}", API_BASE_URL)
    return Response(content=js, media_type="application/javascript", headers={
        "Cache-Control": "no-cache",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET",
    })


# ---------------------------------------------------------------------------
# Event alıcı
# ---------------------------------------------------------------------------

_customer_to_tid: dict[str, dict] = {}
_vid_to_phone: dict[str, str] = {}
_email_to_phone: dict[str, str] = {}
_vid_to_name: dict[str, str] = {}
_email_to_name: dict[str, str] = {}


@router.post("/api/live/event")
async def receive_event(request: Request):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid_json"}, status_code=400)

    tid = str(body.get("tid", "")).strip()
    if not tid:
        return JSONResponse({"ok": False, "error": "missing_tid"}, status_code=400)

    ua = str(body.get("ua", ""))[:512]
    if _is_bot(ua):
        return JSONResponse({"ok": True, "skipped": "bot"})

    # IP adresini erken çek (rate limiting ve event kaydı için)
    raw_ip = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    if not raw_ip:
        try:
            raw_ip = request.client.host if request.client else ""
        except Exception:
            raw_ip = ""

    # Rate limiting — TID başına 500/dk, IP başına 200/dk
    if await store.check_rate_limit(f"rl:tid:{tid}", 500, 60):
        return JSONResponse({"ok": False, "error": "rate_limited"}, status_code=429)
    if raw_ip and await store.check_rate_limit(f"rl:ip:{raw_ip}", 200, 60):
        return JSONResponse({"ok": False, "error": "rate_limited"}, status_code=429)

    vid = str(body.get("vid", ""))[:32]
    event_type = str(body.get("event_type", ""))[:64]
    url = str(body.get("url", ""))[:512]
    if await store.is_duplicate(tid, vid, event_type, url):
        return JSONResponse({"ok": True, "skipped": "duplicate"})

    event = {
        "tid": tid,
        "vid": vid,
        "event_type": event_type,
        "url": url,
        "referrer": str(body.get("referrer", ""))[:256],
        "ts": int(body.get("ts", time.time() * 1000)),
        "ua": ua,
        "sw": int(body.get("sw", 0)),
        "ip": raw_ip[:64] if raw_ip else "",
        "utm": body.get("utm") if isinstance(body.get("utm"), dict) else {},
        "customer_id": str(body.get("customer_id", ""))[:64],
        "data": body.get("data") if isinstance(body.get("data"), dict) else {},
    }

    await store.push_event(tid, event)

    owner = await store.get_tid_owner(tid)

    # Auto-recover: parse old-format TID ({username}_{brand}_{16_hex}) when owner unknown
    if not owner and not tid.startswith("spt_"):
        parts = tid.rsplit("_", 2)
        import re as _re
        if len(parts) == 3 and _re.fullmatch(r"[0-9a-f]{8,32}", parts[2]):
            _u, _b = parts[0], parts[1]
            await store.register_tid_owner(tid, _u, _b)
            owner = (_u, _b)
            # Save pixel_tracking_id back to DB if missing
            try:
                from services.db import get_setting as _gs, set_connection_settings as _scs
                if not _gs(_u, _b, "shopify", "pixel_tracking_id", ""):
                    _scs(_u, _b, "shopify", {"pixel_tracking_id": tid})
                    logger.info("[LIVE] TID auto-recovered and saved to DB: %s", tid[:24])
            except Exception as _e:
                logger.warning("[LIVE] TID DB recovery failed: %s", _e)

    if event.get("customer_id"):
        _customer_to_tid[event["customer_id"]] = {
            "tid": tid, "vid": event["vid"],
            "username": owner[0] if owner else "",
            "brand": owner[1] if owner else "default",
        }

    return JSONResponse({"ok": True})


# ---------------------------------------------------------------------------
# SSE stream
# ---------------------------------------------------------------------------

@router.get("/api/live/stream")
async def sse_stream(request: Request, tid: str = Query(...), current_user: dict = Depends(get_current_user)):
    q = store.subscribe(tid)

    async def generate():
        for ev in await store.get_recent_events(tid, limit=500):
            yield f"data: {json.dumps(ev)}\n\n"
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    ev = await asyncio.wait_for(q.get(), timeout=25.0)
                    yield f"data: {json.dumps(ev)}\n\n"
                except asyncio.TimeoutError:
                    yield ": ping\n\n"
        finally:
            store.unsubscribe(tid, q)

    return StreamingResponse(generate(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
        "Access-Control-Allow-Origin": "*",
    })


@router.get("/api/live/events")
async def get_recent_events(tid: str = Query(...), limit: int = Query(50)):
    limit = min(limit, 5000)
    evs = await store.get_recent_events(tid, limit=limit)
    total = await store.count_events(tid)
    return {"ok": True, "events": evs, "total": total}


# ---------------------------------------------------------------------------
# Pixel kurulum
# ---------------------------------------------------------------------------

def _shopify_headers(token: str) -> dict:
    return {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}


def _shopify_url(domain: str, path: str, version: str = None) -> str:
    domain = domain.replace("https://", "").replace("http://", "").strip().rstrip("/")
    v = version or SHOPIFY_API_VERSION
    return f"https://{domain}/admin/api/{v}/{path}"


def _shopify_graphql(domain: str, token: str, query: str, variables: dict = None, version: str = None) -> dict:
    """Shopify GraphQL Admin API çağrısı. Dönen dict: r.json() içeriği."""
    domain_clean = domain.replace("https://", "").replace("http://", "").strip().rstrip("/")
    v = version or SHOPIFY_API_VERSION
    url = f"https://{domain_clean}/admin/api/{v}/graphql.json"
    r = requests.post(
        url,
        json={"query": query, "variables": variables or {}},
        headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


# REST → GraphQL topic dönüşüm tablosu
_WEBHOOK_TOPIC_MAP = {
    "orders/create":    "ORDERS_CREATE",
    "checkouts/create": "CHECKOUTS_CREATE",
    "checkouts/update": "CHECKOUTS_UPDATE",
    "app/uninstalled":  "APP_UNINSTALLED",
}


def _get_or_create_tid(username: str, brand: str) -> str:
    existing = get_setting(username, brand, "shopify", "pixel_tracking_id", "")
    if existing:
        return existing
    tid = "spt_" + secrets.token_hex(16)
    set_connection_settings(username, brand, "shopify", {"pixel_tracking_id": tid})
    return tid


def _find_our_scripttag(domain: str, token: str, version: str = None) -> Optional[dict]:
    """Mağazada kurulu pixel script tag'ini GraphQL ile bulur."""
    _GQL = """
    query ScriptTags($cursor: String) {
      scriptTags(first: 50, after: $cursor) {
        pageInfo { hasNextPage endCursor }
        edges { node { id src } }
      }
    }
    """
    search_bases = [b for b in [API_BASE_URL, _SHOPIFY_APP_URL] if b]
    cursor = None
    try:
        while True:
            resp = _shopify_graphql(domain, token, _GQL, {"cursor": cursor}, version)
            st = resp.get("data", {}).get("scriptTags", {})
            for edge in st.get("edges", []):
                node = edge["node"]
                src = node.get("src", "")
                if any(src.startswith(base + "/pixel.js") for base in search_bases):
                    return {"id": node["id"], "src": src}
            pi = st.get("pageInfo", {})
            if not pi.get("hasNextPage"):
                break
            cursor = pi.get("endCursor")
    except Exception:
        pass
    return None


@router.get("/api/shopify/pixel/status")
async def pixel_status(username: str = Query(""), brand: str = Query("default")):
    domain = get_setting(username, brand, "shopify", "shop_domain", "")
    token  = get_setting(username, brand, "shopify", "admin_api_token", "")
    tid    = get_setting(username, brand, "shopify", "pixel_tracking_id", "")

    # Primary path: Shopify merchant with Theme App Extension
    # ScriptTag deprecated → TID varlığı = embed konfigüre edilmiş demek.
    # "installed: True" + detected_via="theme_extension" → embed ayarlı ama henüz event yok
    # "installed: True" + detected_via="events" → gerçek trafik var, embed aktif
    if domain and token:
        if not tid:
            # TID DB'de yoksa reverse mapping'den kurtarmayı dene
            tid = await store.get_user_tid(username, brand) or ""
            if tid:
                set_connection_settings(username, brand, "shopify", {"pixel_tracking_id": tid})
                logger.info("[PIXEL] TID reverse mapping'den kurtarıldı: %s", tid[:16])
        if tid:
            await store.register_tid_owner(tid, username, brand)
            # Event gelip gelmediğini kontrol et — TID varlığı değil, gerçek event sayısı
            has_events = (await store.count_events(tid)) > 0
            return {
                "ok": True,
                "installed": True,
                "tracking_id": tid,
                "script_tag_id": None,
                "script_url": None,
                "detected_via": "events" if has_events else "theme_extension",
            }
        return {"ok": True, "installed": False, "tracking_id": None}

    # Fallback: no Shopify credentials (password-login user) — use Redis event presence
    if tid:
        await store.register_tid_owner(tid, username, brand)
        owner = await store.get_tid_owner(tid)
        return {
            "ok": True,
            "installed": bool(owner),
            "tracking_id": tid,
            "script_tag_id": None,
            "script_url": None,
            "detected_via": "events",
        }

    # Last resort: check reverse mapping populated by receive_event auto-recovery
    recovered_tid = await store.get_user_tid(username, brand)
    if recovered_tid:
        logger.info("[PIXEL] TID recovered from reverse mapping: %s", recovered_tid[:24])
        return {
            "ok": True,
            "installed": True,
            "tracking_id": recovered_tid,
            "script_tag_id": None,
            "script_url": None,
            "detected_via": "events",
        }

    return {"ok": True, "installed": False, "tracking_id": None}


@router.post("/api/shopify/pixel/install")
async def pixel_install(username: str = Query(""), brand: str = Query("default")):
    domain  = get_setting(username, brand, "shopify", "shop_domain", "")
    token   = get_setting(username, brand, "shopify", "admin_api_token", "")
    version = get_setting(username, brand, "shopify", "api_version", SHOPIFY_API_VERSION)
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "shopify_not_connected"}, status_code=400)
    existing = _find_our_scripttag(domain, token, version)
    if existing:
        tid = _get_or_create_tid(username, brand)
        await store.register_tid_owner(tid, username, brand)
        return {"ok": True, "already_installed": True, "tracking_id": tid, "script_tag_id": existing["id"]}
    tid = _get_or_create_tid(username, brand)
    await store.register_tid_owner(tid, username, brand)
    script_url = f"{API_BASE_URL}/pixel.js?tid={tid}"
    _MUT = """
    mutation ScriptTagCreate($input: ScriptTagInput!) {
      scriptTagCreate(input: $input) {
        scriptTag { id src }
        userErrors { field message }
      }
    }
    """
    try:
        resp = _shopify_graphql(domain, token, _MUT, {"input": {"event": "ONLOAD", "src": script_url}}, version)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)
    result = resp.get("data", {}).get("scriptTagCreate", {})
    errs = result.get("userErrors", [])
    if errs:
        return JSONResponse({"ok": False, "error": errs[0].get("message")}, status_code=502)
    tag = result.get("scriptTag") or {}
    return {"ok": True, "installed": True, "tracking_id": tid, "script_tag_id": tag.get("id")}


@router.delete("/api/shopify/pixel/uninstall")
async def pixel_uninstall(username: str = Query(""), brand: str = Query("default")):
    domain  = get_setting(username, brand, "shopify", "shop_domain", "")
    token   = get_setting(username, brand, "shopify", "admin_api_token", "")
    version = get_setting(username, brand, "shopify", "api_version", SHOPIFY_API_VERSION)
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "shopify_not_connected"}, status_code=400)
    tag = _find_our_scripttag(domain, token, version)
    if not tag:
        return {"ok": True, "already_uninstalled": True}
    _MUT = """
    mutation ScriptTagDelete($id: ID!) {
      scriptTagDelete(id: $id) {
        deletedScriptTagId
        userErrors { field message }
      }
    }
    """
    try:
        resp = _shopify_graphql(domain, token, _MUT, {"id": tag["id"]}, version)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)
    result = resp.get("data", {}).get("scriptTagDelete", {})
    errs = result.get("userErrors", [])
    if errs:
        return JSONResponse({"ok": False, "error": errs[0].get("message")}, status_code=502)
    return {"ok": True, "uninstalled": True}


# ---------------------------------------------------------------------------
# Müşteri bilgisi
# ---------------------------------------------------------------------------

@router.get("/api/shopify/customer")
async def get_shopify_customer(
    customer_id: str = Query(...),
    username: str = Query(""),
    brand: str = Query("default"),
):
    domain  = get_setting(username, brand, "shopify", "shop_domain", "")
    token   = await store.get_online_token(username, brand) \
              or get_setting(username, brand, "shopify", "admin_api_token", "")
    version = get_setting(username, brand, "shopify", "api_version", SHOPIFY_API_VERSION)
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "shopify_not_connected"}, status_code=400)
    cid = str(customer_id).strip()
    if not cid or not cid.isdigit():
        return JSONResponse({"ok": False, "error": "invalid_customer_id"}, status_code=400)
    try:
        r = requests.get(
            _shopify_url(domain, f"customers/{cid}.json", version),
            headers=_shopify_headers(token), timeout=10,
        )
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)
    if r.status_code == 200:
        c = r.json().get("customer", {})
        return {"ok": True, "customer": {
            "id": c.get("id"), "first_name": c.get("first_name", ""),
            "last_name": c.get("last_name", ""), "email": c.get("email", ""),
            "phone": c.get("phone", ""), "orders_count": c.get("orders_count", 0),
            "total_spent": c.get("total_spent", "0.00"),
        }}
    return JSONResponse({"ok": False, "error": f"shopify_status_{r.status_code}"}, status_code=502)


# ---------------------------------------------------------------------------
# Sipariş Yolculuğu (CustomerJourneySummary via Shopify GraphQL)
# ---------------------------------------------------------------------------

_ORDER_JOURNEY_GQL = """
query OrderJourney($id: ID!) {
  order(id: $id) {
    id
    name
    createdAt
    sourceName
    totalPriceSet { shopMoney { amount currencyCode } }
    channelInformation { channelDefinition { channelName } }
    customerJourneySummary {
      customerOrderIndex
      daysToConversion
      ready
      firstVisit {
        source
        referrerUrl
        occurredAt
        utmParameters { source medium campaign content term }
      }
      lastVisit {
        source
        referrerUrl
        occurredAt
        utmParameters { source medium campaign content term }
      }
      moments(first: 20) {
        nodes {
          ... on CustomerVisit {
            source
            referrerUrl
            occurredAt
            utmParameters { source medium campaign content term }
          }
        }
      }
    }
  }
}
"""

@router.get("/api/shopify/order-journey")
async def get_order_journey(
    order_id: str = Query(...),
    username: str = Query(""),
    brand: str = Query("default"),
    current_user: dict = Depends(get_current_user),
):
    """Shopify GraphQL CustomerJourneySummary — siparişin tam yolculuğunu döner."""
    domain = get_setting(username, brand, "shopify", "shop_domain", "")
    # Online token önce dene (her zaman expiring → 403 hatasını önler)
    token  = await store.get_online_token(username, brand) \
             or get_setting(username, brand, "shopify", "admin_api_token", "")
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "Shopify bağlantısı bulunamadı"}, status_code=400)

    # Numeric ID → GID
    oid = str(order_id).strip()
    gid = oid if oid.startswith("gid://") else f"gid://shopify/Order/{oid}"

    gql_url = f"https://{domain.lstrip('https://').rstrip('/')}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(gql_url, headers=headers,
                                  json={"query": _ORDER_JOURNEY_GQL, "variables": {"id": gid}})
        data = r.json()
    except Exception as exc:
        logger.exception("[JOURNEY] Shopify GraphQL bağlantı hatası")
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)

    if r.status_code != 200:
        return JSONResponse({"ok": False, "error": f"Shopify {r.status_code}: {r.text[:200]}"}, status_code=502)

    errors = data.get("errors")
    if errors:
        first_msg = errors[0].get("message", str(errors)) if isinstance(errors, list) else str(errors)
        logger.warning("[JOURNEY] GQL hata: %s", first_msg)
        return JSONResponse({"ok": False, "error": first_msg}, status_code=502)

    order_data = (data.get("data") or {}).get("order")
    if not order_data:
        return JSONResponse({"ok": False, "error": "Sipariş bulunamadı (GID: " + gid + ")"}, status_code=404)

    # customerJourneySummary null gelirse boş dict ile güvenli dön
    if order_data.get("customerJourneySummary") is None:
        order_data["customerJourneySummary"] = {
            "ready": False, "daysToConversion": None, "customerOrderIndex": None,
            "firstVisit": None, "lastVisit": None, "moments": {"nodes": []},
        }

    return {"ok": True, "order": order_data}


# ---------------------------------------------------------------------------
# Attribution Raporu (siparişlerin kanal/kaynak/kampanya kırılımı → XLSX)
# ---------------------------------------------------------------------------

_ATTR_REPORT_GQL = """
query AttrReport($cursor: String, $query: String!) {
  orders(first: 50, after: $cursor, query: $query, sortKey: CREATED_AT, reverse: true) {
    pageInfo { hasNextPage endCursor }
    nodes {
      id
      name
      createdAt
      sourceName
      test
      displayFinancialStatus
      totalPriceSet { shopMoney { amount currencyCode } }
      customer { displayName }
      channelInformation { channelDefinition { channelName } }
      customerJourneySummary {
        ready
        daysToConversion
        customerOrderIndex
        momentsCount { count }
        firstVisit { source referrerUrl occurredAt utmParameters { source medium campaign } }
        lastVisit  { source referrerUrl occurredAt utmParameters { source medium campaign } }
      }
    }
  }
}
"""

_ATTR_HEADERS = {
    "tr": ["Sipariş", "Tarih", "Müşteri", "Yeni/Tekrar", "Tutar", "Para Birimi",
           "Ödeme Durumu", "WhatsApp", "Satış Kanalı",
           "İlk Kaynak", "İlk Kampanya", "Son Kaynak", "Son Medium", "Son Kampanya",
           "Yönlendiren (Referrer)", "İlk Ziyaret", "Son Ziyaret", "Dönüşüm Günü",
           "Temas Noktası", "Sipariş Sırası"],
    "en": ["Order", "Date", "Customer", "New/Returning", "Amount", "Currency",
           "Payment Status", "WhatsApp", "Sales Channel",
           "First Source", "First Campaign", "Last Source", "Last Medium", "Last Campaign",
           "Referrer", "First Visit", "Last Visit", "Days to Conversion",
           "Touchpoints", "Order Index"],
}

# Özet sayfası başlıkları
_ATTR_SUMMARY_TXT = {
    "tr": {"sheet": "Özet", "by_last": "Son Dokunuş — Kaynak Bazlı",
           "by_first": "İlk Dokunuş — Kaynak Bazlı", "wa": "WhatsApp Attribution",
           "cols": ["Kaynak", "Sipariş", "Ciro", "Ort. Sepet (AOV)", "Ciro %"],
           "total": "TOPLAM", "unknown": "(bilinmiyor)",
           "wa_rows": ["WhatsApp'tan gelen sipariş", "WhatsApp cirosu",
                       "Toplam sipariş", "Toplam ciro", "WhatsApp ciro payı"],
           "note": "Not: İptal/iade siparişler dahildir; test siparişleri hariç tutulmuştur."},
    "en": {"sheet": "Summary", "by_last": "Last-touch — by Source",
           "by_first": "First-touch — by Source", "wa": "WhatsApp Attribution",
           "cols": ["Source", "Orders", "Revenue", "Avg Order (AOV)", "Revenue %"],
           "total": "TOTAL", "unknown": "(unknown)",
           "wa_rows": ["WhatsApp-attributed orders", "WhatsApp revenue",
                       "Total orders", "Total revenue", "WhatsApp revenue share"],
           "note": "Note: cancelled/refunded orders are included; test orders are excluded."},
}


def _attr_visit_fields(visit: dict) -> tuple:
    """firstVisit/lastVisit dict'inden (source, medium, campaign, referrer, occurredAt) çıkarır."""
    if not visit:
        return ("", "", "", "", "")
    utm = visit.get("utmParameters") or {}
    return (
        visit.get("source") or "",
        utm.get("medium") or "",
        utm.get("campaign") or "",
        visit.get("referrerUrl") or "",
        visit.get("occurredAt") or "",
    )


@router.get("/api/shopify/attribution-report")
async def get_attribution_report(
    username: str = Query(""),
    brand: str = Query("default"),
    days: int = Query(30),
    lang: str = Query("tr"),
    current_user: dict = Depends(get_current_user),
):
    """Son N gündeki siparişlerin kanal/kaynak/kampanya attribution kırılımını XLSX olarak indirir."""
    from datetime import datetime, timezone, timedelta
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    days = max(1, min(int(days), 365))
    lang = lang if lang in ("tr", "en") else "tr"
    YENI  = "Yeni" if lang == "tr" else "New"
    TEKRAR = "Tekrar" if lang == "tr" else "Returning"
    EVET  = "Evet ✓" if lang == "tr" else "Yes ✓"

    domain = get_setting(username, brand, "shopify", "shop_domain", "")
    token  = await store.get_online_token(username, brand) \
             or get_setting(username, brand, "shopify", "admin_api_token", "")
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "shopify_not_connected"}, status_code=400)

    # WhatsApp-attributed sipariş kimlikleri (kendi attribution verimiz)
    wa_ids = set()
    try:
        for co in await store.get_converted_orders(username, brand, limit=200):
            if co.get("wa_attributed"):
                if co.get("order_id"):
                    wa_ids.add(str(co["order_id"]))
                if co.get("order_number"):
                    wa_ids.add(str(co["order_number"]).lstrip("#"))
    except Exception:
        logger.warning("[ATTR-REPORT] WA attribution verisi okunamadı")

    since = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    gql_query = f"created_at:>={since}"

    rows = []
    agg = []   # (last_src, first_src, amount_float, is_wa) — özet için
    cursor = None
    pages = 0
    try:
        while True:
            data = await asyncio.to_thread(
                _shopify_graphql, domain, token, _ATTR_REPORT_GQL,
                {"cursor": cursor, "query": gql_query},
            )
            errs = data.get("errors")
            if errs:
                msg = errs[0].get("message", str(errs)) if isinstance(errs, list) else str(errs)
                logger.warning("[ATTR-REPORT] GQL hata: %s", msg)
                return JSONResponse({"ok": False, "error": msg}, status_code=502)

            conn = ((data.get("data") or {}).get("orders") or {})
            for o in conn.get("nodes", []):
                if o.get("test"):
                    continue   # test siparişlerini ayıkla
                money = (o.get("totalPriceSet") or {}).get("shopMoney") or {}
                cust  = o.get("customer") or {}
                chan  = ((o.get("channelInformation") or {}).get("channelDefinition") or {})
                cjs   = o.get("customerJourneySummary") or {}
                f_src, _f_med, f_camp, _f_ref, f_occ = _attr_visit_fields(cjs.get("firstVisit"))
                l_src, l_med, l_camp, l_ref, l_occ   = _attr_visit_fields(cjs.get("lastVisit"))
                try:
                    amount = float(money.get("amount") or 0)
                    amount_f = amount
                except (TypeError, ValueError):
                    amount = money.get("amount") or ""
                    amount_f = 0.0

                # Yeni / Tekrar müşteri
                idx = cjs.get("customerOrderIndex")
                if idx == 1:
                    new_ret = YENI
                elif isinstance(idx, int) and idx > 1:
                    new_ret = TEKRAR
                else:
                    new_ret = ""

                # WhatsApp eşleşmesi (numeric id ya da sipariş no)
                oid_num = str(o.get("id") or "").rsplit("/", 1)[-1]
                oname   = str(o.get("name") or "").lstrip("#")
                is_wa = bool(wa_ids) and (oid_num in wa_ids or oname in wa_ids)

                rows.append([
                    o.get("name") or "",
                    (o.get("createdAt") or "")[:10],
                    cust.get("displayName") or "",
                    new_ret,
                    amount,
                    money.get("currencyCode") or "",
                    o.get("displayFinancialStatus") or "",
                    EVET if is_wa else "",
                    chan.get("channelName") or o.get("sourceName") or "",
                    f_src,
                    f_camp,
                    l_src,
                    l_med,
                    l_camp,
                    l_ref,
                    (f_occ or "")[:10],
                    (l_occ or "")[:10],
                    cjs.get("daysToConversion") if cjs.get("daysToConversion") is not None else "",
                    (cjs.get("momentsCount") or {}).get("count", "") if cjs.get("momentsCount") else "",
                    idx if idx is not None else "",
                ])
                agg.append((l_src, f_src, amount_f, is_wa))

            page_info = conn.get("pageInfo") or {}
            pages += 1
            if page_info.get("hasNextPage") and pages < 40:   # güvenlik tavanı: ~2000 sipariş
                cursor = page_info.get("endCursor")
            else:
                break
    except Exception as exc:
        logger.exception("[ATTR-REPORT] Shopify bağlantı hatası")
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)

    # ── XLSX oluştur ──
    headers = _ATTR_HEADERS[lang]
    txt = _ATTR_SUMMARY_TXT[lang]
    wb = Workbook()

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    sec_font  = Font(bold=True, size=12, color="111827")
    col_font  = Font(bold=True, size=10, color="374151")
    col_fill  = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    tot_font  = Font(bold=True, size=10)
    center    = Alignment(horizontal="center", vertical="center")

    # ── Sayfa 1: Özet ──
    sm = wb.active
    sm.title = txt["sheet"]
    currency = next((r[5] for r in rows if r[5]), "")
    total_orders  = len(agg)
    total_revenue = sum(a[2] for a in agg)

    def _breakdown(key_idx):
        d = {}
        for a in agg:
            k = a[key_idx] or txt["unknown"]
            cnt, rev = d.get(k, (0, 0.0))
            d[k] = (cnt + 1, rev + a[2])
        return sorted(d.items(), key=lambda kv: kv[1][1], reverse=True)

    def _write_table(start_row, title, data):
        sm.cell(row=start_row, column=1, value=title).font = sec_font
        hr = start_row + 1
        for ci, ch in enumerate(txt["cols"], start=1):
            c = sm.cell(row=hr, column=ci, value=ch)
            c.font = col_font
            c.fill = col_fill
            c.alignment = center
        r = hr + 1
        for src, (cnt, rev) in data:
            share = (rev / total_revenue * 100) if total_revenue else 0
            sm.cell(row=r, column=1, value=src)
            sm.cell(row=r, column=2, value=cnt)
            sm.cell(row=r, column=3, value=round(rev, 2))
            sm.cell(row=r, column=4, value=round(rev / cnt, 2) if cnt else 0)
            sm.cell(row=r, column=5, value=f"{share:.1f}%")
            r += 1
        # toplam satırı
        sm.cell(row=r, column=1, value=txt["total"]).font = tot_font
        sm.cell(row=r, column=2, value=total_orders).font = tot_font
        sm.cell(row=r, column=3, value=round(total_revenue, 2)).font = tot_font
        sm.cell(row=r, column=4, value=round(total_revenue / total_orders, 2) if total_orders else 0).font = tot_font
        sm.cell(row=r, column=5, value="100%").font = tot_font
        return r + 2   # sonraki tablo için boş satır

    nxt = _write_table(1, f"{txt['by_last']}  ({currency})", _breakdown(0))
    nxt = _write_table(nxt, f"{txt['by_first']}  ({currency})", _breakdown(1))

    # WhatsApp özeti
    wa_orders  = sum(1 for a in agg if a[3])
    wa_revenue = sum(a[2] for a in agg if a[3])
    sm.cell(row=nxt, column=1, value=txt["wa"]).font = sec_font
    wr = nxt + 1
    wa_vals = [wa_orders, round(wa_revenue, 2), total_orders, round(total_revenue, 2),
               f"{(wa_revenue / total_revenue * 100):.1f}%" if total_revenue else "0%"]
    for label, val in zip(txt["wa_rows"], wa_vals):
        sm.cell(row=wr, column=1, value=label).font = col_font
        sm.cell(row=wr, column=2, value=val)
        wr += 1
    sm.cell(row=wr + 1, column=1, value=txt["note"]).font = Font(italic=True, size=9, color="6B7280")

    for col, w in zip("ABCDE", [30, 12, 16, 18, 12]):
        sm.column_dimensions[col].width = w

    # ── Sayfa 2: Detay ──
    ws = wb.create_sheet("Attribution" if lang == "en" else "Detay")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    for r in rows:
        ws.append(r)

    widths = [12, 12, 22, 12, 11, 8, 14, 11, 18, 16, 28, 16, 12, 28, 30, 12, 12, 10, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"attribution-report-{since}-{days}d.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Report-Rows": str(len(rows)),
        },
    )


# ---------------------------------------------------------------------------
# RFM Segmentasyon endpoint'i
# ---------------------------------------------------------------------------

_RFM_ORDERS_GQL = """
query RFMOrders($cursor: String, $query: String!) {
  orders(first: 250, after: $cursor, query: $query, sortKey: CREATED_AT) {
    pageInfo { hasNextPage endCursor }
    nodes {
      createdAt
      totalPriceSet { shopMoney { amount currencyCode } }
      customer {
        id
        displayName
        email
      }
    }
  }
}
"""


@router.get("/api/shopify/customers/rfm")
async def get_rfm_segments(
    username: str = Query(""),
    brand: str = Query("default"),
    days: int = Query(90),
    current_user: dict = Depends(get_current_user),
):
    """Son N günün siparişlerinden müşterileri RFM ile segmentler."""
    import bisect, math
    from datetime import datetime, timezone, timedelta

    domain = get_setting(username, brand, "shopify", "shop_domain", "")
    token  = await store.get_online_token(username, brand) \
             or get_setting(username, brand, "shopify", "admin_api_token", "")
    logger.info("[RFM] domain=%r token_set=%s user=%s brand=%s", domain or "(empty)", bool(token), username[:20], brand)
    if not domain or not token:
        # Diagnoz: raw payload'u logla
        try:
            from services.db import _get_conn as _dbc
            import psycopg2.extras as _pex
            with _dbc() as _c:
                with _c.cursor(cursor_factory=_pex.RealDictCursor) as _cur:
                    _cur.execute("SELECT payload_json FROM integration_connections WHERE username=%s AND brand=%s AND integration_id='shopify' LIMIT 1", (username, brand))
                    row = _cur.fetchone()
                    if row:
                        raw = row["payload_json"]
                        if isinstance(raw, str):
                            raw = json.loads(raw)
                        keys = list(raw.keys()) if isinstance(raw, dict) else []
                        domain_val = raw.get("shop_domain", "(key missing)") if isinstance(raw, dict) else "N/A"
                        token_val = raw.get("admin_api_token", "(key missing)") if isinstance(raw, dict) else "N/A"
                        logger.warning("[RFM] DB row keys: %s | shop_domain=%r | token=%s",
                            keys, domain_val, "SET" if token_val and token_val != "(key missing)" else repr(token_val))
                    else:
                        logger.warning("[RFM] DB row NOT FOUND for user=%s brand=%s", username, brand)
        except Exception as _de:
            logger.warning("[RFM] DB diagnoz hatası: %s", _de)
        return JSONResponse({"ok": False, "error": "Shopify bağlantısı bulunamadı — admin API token gerekli"}, status_code=400)

    since = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
    gql_url = f"https://{domain.lstrip('https://').rstrip('/')}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
    q_filter = f"created_at:>{since}"

    # Sipariş topla (maks 2 sayfa = 500 sipariş)
    all_nodes = []
    cursor = None
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            for _ in range(2):
                payload = {"query": _RFM_ORDERS_GQL, "variables": {"cursor": cursor, "query": q_filter}}
                r = await client.post(gql_url, headers=headers, json=payload)
                if r.status_code != 200:
                    break
                data = r.json()
                orders_data = ((data.get("data") or {}).get("orders") or {})
                all_nodes.extend(orders_data.get("nodes") or [])
                pi = orders_data.get("pageInfo") or {}
                if not pi.get("hasNextPage"):
                    break
                cursor = pi.get("endCursor")
    except Exception as exc:
        logger.exception("[RFM] Shopify bağlantı hatası")
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)

    # Müşteri bazında gruplama
    now_dt = datetime.now(timezone.utc)
    customers: dict[str, dict] = {}
    currency = "TRY"
    for node in all_nodes:
        cust = node.get("customer") or {}
        cid = cust.get("id", "")
        if not cid:
            continue
        amt = float((node.get("totalPriceSet") or {}).get("shopMoney", {}).get("amount") or 0)
        currency = (node.get("totalPriceSet") or {}).get("shopMoney", {}).get("currencyCode", currency)
        try:
            order_dt = datetime.fromisoformat(node["createdAt"].replace("Z", "+00:00"))
        except Exception:
            continue
        r_days = (now_dt - order_dt).days
        if cid not in customers:
            customers[cid] = {
                "id": cid,
                "name": cust.get("displayName", ""),
                "email": cust.get("email", ""),
                "r_days": r_days,
                "frequency": 0,
                "monetary": 0.0,
            }
        c = customers[cid]
        c["frequency"] += 1
        c["monetary"] += amt
        if r_days < c["r_days"]:
            c["r_days"] = r_days

    if not customers:
        return {"ok": True, "segments": {}, "customers": [], "currency": currency, "order_count": 0}

    clist = list(customers.values())

    # Quintile sınırları
    def quintiles(vals):
        s = sorted(vals)
        n = len(s)
        return [s[max(0, int(n * p / 5) - 1)] for p in range(1, 5)]

    r_q = quintiles([c["r_days"] for c in clist])
    f_q = quintiles([c["frequency"] for c in clist])
    m_q = quintiles([c["monetary"] for c in clist])

    def score(val, quint, invert=False):
        sc = bisect.bisect_right(quint, val) + 1  # 1-5
        return 6 - sc if invert else sc

    def segment(r, f):
        if r >= 4 and f >= 4: return "champions"
        if r >= 3 and f >= 3: return "loyal"
        if r >= 4 and f < 3:  return "promising"
        if r == 1 and f >= 2: return "lost"
        if r <= 2 and f >= 3: return "at_risk"
        if r >= 4 and f == 1: return "new"
        return "needs_attention"

    for c in clist:
        r = score(c["r_days"], r_q, invert=True)  # lower days = higher R
        f = score(c["frequency"], f_q)
        m = score(c["monetary"], m_q)
        c["r_score"] = r
        c["f_score"] = f
        c["m_score"] = m
        c["rfm_score"] = r * 100 + f * 10 + m
        c["segment"] = segment(r, f)
        c["monetary"] = round(c["monetary"], 2)

    # Segment sayımları
    seg_counts: dict[str, int] = {}
    for c in clist:
        seg_counts[c["segment"]] = seg_counts.get(c["segment"], 0) + 1

    # Her segmentten en fazla 10 müşteri (monetary'e göre sıralı)
    seg_customers: dict[str, list] = {}
    for seg in set(c["segment"] for c in clist):
        top = sorted([c for c in clist if c["segment"] == seg], key=lambda x: -x["monetary"])[:10]
        seg_customers[seg] = top

    return {
        "ok": True,
        "segments": seg_counts,
        "seg_customers": seg_customers,
        "total_customers": len(clist),
        "order_count": len(all_nodes),
        "currency": currency,
        "days": days,
    }


# ---------------------------------------------------------------------------
# Ürün stok endpoint'i — AdProductGrid widget için
# ---------------------------------------------------------------------------

_PRODUCTS_STOCK_GQL = """
query ProductsStock($query: String!) {
  products(first: 30, query: $query) {
    nodes {
      handle
      title
      totalInventory
      tracksInventory
      featuredImage { url }
      variants(first: 1) {
        nodes { inventoryQuantity }
      }
    }
  }
}
"""


@router.get("/api/shopify/products/stock")
async def get_products_stock(
    username: str = Query(""),
    brand: str = Query("default"),
    handles: str = Query(""),
    current_user: dict = Depends(get_current_user),
):
    """Virgülle ayrılmış ürün handle listesi için stok bilgisi döner."""
    domain = get_setting(username, brand, "shopify", "shop_domain", "")
    token  = await store.get_online_token(username, brand) \
             or get_setting(username, brand, "shopify", "admin_api_token", "")
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "shopify_not_connected"}, status_code=400)

    handle_list = [h.strip() for h in handles.split(",") if h.strip()][:30]
    if not handle_list:
        return {"ok": True, "products": {}}

    gql_url = f"https://{domain.lstrip('https://').rstrip('/')}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
    handle_filter = " OR ".join(f"handle:{h}" for h in handle_list)

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(gql_url, headers=headers,
                                  json={"query": _PRODUCTS_STOCK_GQL, "variables": {"query": handle_filter}})
            if r.status_code != 200:
                return JSONResponse({"ok": False, "error": f"Shopify {r.status_code}"}, status_code=502)
            data = r.json()
            nodes = ((data.get("data") or {}).get("products") or {}).get("nodes", [])
            products = {}
            for node in nodes:
                h = node.get("handle", "")
                if not h:
                    continue
                products[h] = {
                    "title": node.get("title", ""),
                    "available": node.get("totalInventory") if node.get("tracksInventory") else None,
                    "image": (node.get("featuredImage") or {}).get("url", ""),
                }
            return {"ok": True, "products": products}
    except Exception as exc:
        logger.exception("[STOCK] Shopify bağlantı hatası")
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)


# ---------------------------------------------------------------------------
# Webhook endpoint'leri
# ---------------------------------------------------------------------------

async def _verify_webhook(request: Request, token: str, username: str, brand: str) -> Optional[bytes]:
    """Shopify webhook doğrulama — ÖNCE HMAC imzası (asıl güvenlik), sonra URL token.
    HMAC, webhook URL'sindeki token zamanla değişse/sürüklense bile çalışır
    (eski kayıtlı webhook'lar bu yüzden 401 alıyordu). Geçerliyse raw body döner,
    değilse None."""
    body = await request.body()
    # 1) Shopify HMAC imzası — en güvenilir doğrulama
    hmac_header = request.headers.get("x-shopify-hmac-sha256", "")
    if hmac_header:
        try:
            from routers.gdpr import _verify_hmac
            if _verify_hmac(body, hmac_header):
                return body
        except Exception as e:
            logger.warning("[WEBHOOK] HMAC doğrulama hatası: %s", e)
    # 2) Geriye dönük: URL token eşleşmesi (HMAC yoksa/secret yoksa)
    stored_token = get_setting(username, brand, "shopify", "webhook_token", "")
    if not stored_token or token == stored_token:
        return body
    # 3) Shop-domain eşleşmesi — eski app secret'ı ile imzalanmış / token sürüklenmiş
    #    webhook'lar için güvenli fallback. Shopify her webhook'ta bu başlığı gönderir;
    #    merchant'ın kayıtlı mağaza adresiyle birebir eşleşiyorsa kabul edilir.
    shop_hdr = (request.headers.get("x-shopify-shop-domain", "") or "").strip().lower()
    if shop_hdr:
        stored_domain = get_setting(username, brand, "shopify", "shop_domain", "")
        sd = stored_domain.replace("https://", "").replace("http://", "").rstrip("/").lower()
        if sd and shop_hdr == sd:
            logger.info("[WEBHOOK] shop-domain ile kabul (HMAC/token drift): %s", shop_hdr)
            return body
    logger.warning("[WEBHOOK] yetkisiz: %s:%s hmac=%s token_match=False shop_hdr=%s",
                   username, brand, bool(hmac_header), shop_hdr or "-")
    return None


@router.post("/api/shopify/webhook/orders-create")
async def shopify_orders_webhook(
    request: Request,
    token: str = Query(""),
    username: str = Query(""),
    brand: str = Query("default"),
):
    _body = await _verify_webhook(request, token, username, brand)
    if _body is None:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    try:
        order = json.loads(_body)
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid_json"}, status_code=400)

    customer     = order.get("customer") or {}
    customer_id  = str(customer.get("id", "")).strip()
    order_number = str(order.get("order_number") or order.get("name") or "")
    total_price  = str(order.get("total_price", "0"))

    checkout_token = str(order.get("checkout_token") or "").strip()
    if checkout_token:
        await store.mark_checkout_completed(checkout_token)
        await store.mark_flow_converted(username, brand, checkout_token)

        # WA dönüşüm sipariş detaylarını kaydet
        co_data = await store.get_checkout(checkout_token) or {}
        landing = order.get("landing_site") or ""
        channel = "Direct"
        try:
            from urllib.parse import urlparse, parse_qs as _parse_qs
            _qs = _parse_qs(urlparse(landing).query)
            _src = _qs.get("utm_source", [""])[0]
            _med = _qs.get("utm_medium", [""])[0]
            if _src:
                channel = f"{_src} / {_med}" if _med else _src
        except Exception:
            pass
        if channel == "Direct":
            _sname = order.get("source_name", "")
            if _sname and _sname not in ("web", ""):
                channel = _sname
        _cname = f"{customer.get('first_name','').strip()} {customer.get('last_name','').strip()}".strip()
        _items = []
        for _it in (order.get("line_items") or [])[:5]:
            _items.append({"title": _it.get("title",""), "quantity": _it.get("quantity",1), "price": str(_it.get("price","0"))})
        # WA attribution — was a WA message sent for this checkout?
        wa_attributed = await store.is_wa_sent(checkout_token)
        await store.save_converted_order(username, brand, {
            "order_id":      str(order.get("id", "")),
            "order_number":  str(order.get("order_number") or order.get("name") or ""),
            "total_price":   str(order.get("total_price", "0")),
            "currency":      order.get("currency", "TRY"),
            "customer_name": _cname or co_data.get("name", ""),
            "phone":         co_data.get("phone", ""),
            "product":       co_data.get("product", ""),
            "line_items":    _items,
            "channel":       channel,
            "vid":           co_data.get("vid", ""),
            "wa_attributed": wa_attributed,
            "ts":            int(time.time() * 1000),
        })

    # Sipariş sonrası WA gönder (ayarlarda post_order aktifse)
    settings = await store.get_flow_settings(username, brand)
    post_order = settings.get("post_order") or {}
    if post_order.get("enabled") and settings.get("wa_token") and settings.get("phone_number_id"):
        phone = ""
        phone_src = order.get("billing_address") or order.get("shipping_address") or {}
        phone = str(order.get("phone") or customer.get("phone") or phone_src.get("phone") or "").strip()
        if phone:
            digits = "".join(c for c in phone if c.isdigit())
            if digits.startswith("0"):
                phone = f"+9{digits}"
            elif not phone.startswith("+"):
                phone = f"+{digits}"
            from services.wa_sender import send_wa_template
            customer_name = f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()
            tmpl = post_order.get("template", "siparis_onay")
            await send_wa_template(
                settings["wa_token"], settings["phone_number_id"], phone,
                name=customer_name, order_number=str(order.get("name") or order.get("order_number") or ""),
                template_name=tmpl,
            )

    session_info = _customer_to_tid.get(customer_id) if customer_id else None

    line_items = []
    for item in (order.get("line_items") or [])[:15]:
        line_items.append({
            "title": item.get("title", ""), "quantity": item.get("quantity", 1),
            "price": str(item.get("price", "0")), "variant_id": str(item.get("variant_id", "")),
            "sku": item.get("sku", ""),
        })

    ev_data = {
        "order_id": str(order.get("id", "")), "order_number": order_number,
        "total_price": total_price, "currency": order.get("currency", "TRY"),
        "customer_name": f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip(),
        "line_items": line_items,
    }

    if session_info:
        tid = session_info["tid"]
        vid = session_info["vid"]
        ev = {
            "tid": tid, "vid": vid, "event_type": "checkout_completed",
            "url": "", "referrer": "", "ts": int(time.time() * 1000),
            "ua": "", "sw": 0, "ip": "", "utm": {}, "customer_id": customer_id,
            "data": ev_data,
        }
        await store.push_event(tid, ev)
        return JSONResponse({"ok": True, "matched": True, "tid": tid})

    # Fallback: checkout webhook stored vid+tid — push event via that path
    if checkout_token and co_data.get("vid") and co_data.get("tid"):
        tid = co_data["tid"]
        vid = co_data["vid"]
        ev = {
            "tid": tid, "vid": vid, "event_type": "checkout_completed",
            "url": "", "referrer": "", "ts": int(time.time() * 1000),
            "ua": "", "sw": 0, "ip": "", "utm": {}, "customer_id": customer_id,
            "data": ev_data,
        }
        await store.push_event(tid, ev)
        return JSONResponse({"ok": True, "matched": True, "via": "checkout_data", "tid": tid})

    return JSONResponse({"ok": True, "matched": False})


@router.post("/api/shopify/webhook/checkouts-create")
async def shopify_checkouts_webhook(
    request: Request,
    token: str = Query(""),
    username: str = Query(""),
    brand: str = Query("default"),
):
    _body = await _verify_webhook(request, token, username, brand)
    if _body is None:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    try:
        checkout = json.loads(_body)
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid_json"}, status_code=400)

    # ÖNEMLİ: yalnızca siparişi verenin telefonu (iletişim → fatura). shipping_address.phone
    # KASTEN kullanılmaz — o, ürünün gönderileceği kişi (ör. hediye alan arkadaş) olabilir;
    # alakasız birine "sepetinde ürün var" mesajı gitmesin diye. Veren telefonu yoksa atlanır.
    phone       = str(checkout.get("phone") or checkout.get("billing_address", {}).get("phone") or "").strip()
    email       = str(checkout.get("email") or "").strip().lower()
    customer    = checkout.get("customer") or {}
    customer_id = str(customer.get("id") or "").strip()
    first_name  = str(customer.get("first_name") or checkout.get("billing_address", {}).get("first_name") or "").strip()
    last_name   = str(customer.get("last_name")  or checkout.get("billing_address", {}).get("last_name")  or "").strip()
    customer_name = f"{first_name} {last_name}".strip()

    # Analitik indeksi — TÜM checkout'lar (telefonsuz dahil) CHECKOUT/ABANDONED kartları için.
    # Telefon kontrolünden ÖNCE yapılır ki gerçek checkout sayısı eksiksiz olsun.
    _co_token = str(checkout.get("token") or checkout.get("id") or "").strip()
    if _co_token:
        _co_lines = checkout.get("line_items") or []
        await store.index_checkout(username, brand, _co_token, int(time.time() * 1000), {
            "customer_name": customer_name,
            "product": _co_lines[0].get("title", "") if _co_lines else "",
            "line_items": [{"title": li.get("title", ""), "quantity": li.get("quantity", 1)} for li in _co_lines[:5]],
            "total_price": str(checkout.get("total_price") or checkout.get("subtotal_price") or "0"),
        })

    if not phone:
        logger.info("[CHECKOUT] telefon yok — indekslendi, WA atlandı email=%s", email or "-")
        return JSONResponse({"ok": True, "matched": False, "reason": "no_phone"})

    if not phone.startswith("+"):
        digits = "".join(c for c in phone if c.isdigit())
        phone = f"+9{digits}" if digits.startswith("0") else f"+90{digits}"

    matched_vid = None
    if customer_id and customer_id in _customer_to_tid:
        matched_vid = _customer_to_tid[customer_id].get("vid")

    if matched_vid:
        _vid_to_phone[matched_vid] = phone
        if customer_name:
            _vid_to_name[matched_vid] = customer_name
    if email:
        _email_to_phone[email] = phone
        if customer_name:
            _email_to_name[email] = customer_name

    checkout_token = str(checkout.get("token") or checkout.get("id") or "").strip()
    line_items = checkout.get("line_items") or []
    product = line_items[0].get("title", "") if line_items else ""
    # variant_id de saklanır → cihazlar arası sepet kurtarma (cart permalink) için gerekli
    co_items = [{"title": li.get("title", ""), "quantity": li.get("quantity", 1),
                 "variant_id": str(li.get("variant_id", ""))} for li in line_items[:5]]
    if checkout_token and phone:
        pixel_tid = get_setting(username, brand, "shopify", "pixel_tracking_id", "")
        total_price = str(checkout.get("total_price") or checkout.get("subtotal_price") or "0")
        await store.save_checkout(checkout_token, {
            "token": checkout_token,
            "phone": phone,
            "name": customer_name,
            "product": product,
            "line_items": co_items,
            "total_price": total_price,
            "username": username,
            "brand": brand,
            "vid": matched_vid or "",
            "tid": pixel_tid or "",
            "ts": int(time.time() * 1000),
        })
        if customer_name:
            await store.set_phone_name(phone, customer_name)
        logger.info("[CHECKOUT] kaydedildi phone=***%s name=%s product=%s", phone[-4:], customer_name or "-", product[:30] or "-")
    else:
        logger.info("[CHECKOUT] token veya telefon eksik token=%s phone=%s", bool(checkout_token), bool(phone))

    return JSONResponse({"ok": True, "matched": bool(matched_vid), "vid": matched_vid, "checkout_saved": bool(checkout_token and phone)})


@router.post("/api/shopify/webhook/register")
async def register_order_webhook(username: str = Query(""), brand: str = Query("default")):
    domain  = get_setting(username, brand, "shopify", "shop_domain", "")
    # Shopify non-expiring offline token'ları reddediyor → EXPIRING (online) token tercih et.
    token   = await store.get_online_token(username, brand) \
              or get_setting(username, brand, "shopify", "admin_api_token", "")
    # Webhook kaydı GÜNCEL sürümle yapılmalı — bazı mağazalarda kayıtlı api_version
    # eski ("2024-10") kalıp Shopify'dan 403 aldırıyordu.
    version = SHOPIFY_API_VERSION
    if not domain or not token:
        return JSONResponse({"ok": False, "error": "shopify_not_connected"}, status_code=400)

    wh_token = get_setting(username, brand, "shopify", "webhook_token", "")
    if not wh_token:
        wh_token = secrets.token_hex(16)
        set_connection_settings(username, brand, "shopify", {"webhook_token": wh_token})

    topics = [
        ("orders/create",    "orders-create"),
        ("checkouts/create", "checkouts-create"),
        ("checkouts/update", "checkouts-create"),
    ]
    _MUT = """
    mutation WebhookCreate($topic: WebhookSubscriptionTopic!, $callbackUrl: URL!) {
      webhookSubscriptionCreate(topic: $topic, webhookSubscription: {callbackUrl: $callbackUrl, format: JSON}) {
        webhookSubscription { id }
        userErrors { field message }
      }
    }
    """
    results = {}
    for topic, slug in topics:
        callback_url = (
            f"{API_BASE_URL}/api/shopify/webhook/{slug}"
            f"?token={wh_token}&username={username}&brand={brand}"
        )
        gql_topic = _WEBHOOK_TOPIC_MAP.get(topic, topic.upper().replace("/", "_"))
        try:
            resp = _shopify_graphql(domain, token, _MUT, {"topic": gql_topic, "callbackUrl": callback_url}, version)
            result = resp.get("data", {}).get("webhookSubscriptionCreate", {})
            errs = result.get("userErrors", [])
            if errs and "already" in (errs[0].get("message") or "").lower():
                results[topic] = {"ok": True, "already_registered": True}
            elif errs:
                results[topic] = {"ok": False, "error": errs[0].get("message")}
                logger.warning("[WH-REGISTER] userError topic=%s errs=%s", topic, errs)
            else:
                wh = result.get("webhookSubscription") or {}
                results[topic] = {"ok": True, "webhook_id": wh.get("id")}
                logger.info("[WH-REGISTER] ✓ %s → %s", topic, wh.get("id"))
        except Exception as e:
            _txt = ""
            _r = getattr(e, "response", None)
            if _r is not None:
                try: _txt = _r.text[:400]
                except Exception: pass
            logger.warning("[WH-REGISTER] hata topic=%s ver=%s err=%s body=%s", topic, version, e, _txt)
            results[topic] = {"ok": False, "error": str(e), "detail": _txt}

    set_connection_settings(username, brand, "shopify", {"webhooks_registered": "1"})
    return {"ok": True, "results": results}


@router.get("/api/shopify/webhook/status")
async def get_webhook_status(username: str = Query(""), brand: str = Query("default")):
    domain  = get_setting(username, brand, "shopify", "shop_domain", "")
    token   = await store.get_online_token(username, brand) \
              or get_setting(username, brand, "shopify", "admin_api_token", "")
    version = SHOPIFY_API_VERSION
    if not domain or not token:
        return {"ok": False, "webhooks": []}
    _QRY = """
    query WebhookStatus {
      webhookSubscriptions(first: 20, topics: ORDERS_CREATE) {
        edges { node { id callbackUrl topic } }
      }
    }
    """
    try:
        resp = _shopify_graphql(domain, token, _QRY, version=version)
        edges = resp.get("data", {}).get("webhookSubscriptions", {}).get("edges", [])
        hooks = [e["node"] for e in edges]
        ours  = [h for h in hooks if API_BASE_URL in h.get("callbackUrl", "")]
        return {"ok": True, "registered": bool(ours), "webhooks": ours}
    except Exception:
        pass
    return {"ok": False, "webhooks": []}


# ── Checkout istatistikleri (CHECKOUT/ABANDONED kartları — gerçek webhook verisi) ──

@router.get("/api/live/checkout-stats")
async def checkout_stats_endpoint(
    username: str = Query(""),
    brand: str = Query("default"),
    current_user: dict = Depends(get_current_user),
):
    """Bugün başlatılan/tamamlanan/terk edilen checkout sayıları + listeler."""
    now_tr = time.time() + 3 * 3600
    start_of_day_tr = now_tr - (now_tr % 86400)
    start_ms = int((start_of_day_tr - 3 * 3600) * 1000)
    stats = await store.get_checkout_stats(username, brand, start_ms)
    return {"ok": True, **stats}


# ── Meta WA gelen mesaj webhook'u (opt-out) ───────────────────────────────────

@router.get("/api/wa/webhook")
async def wa_webhook_verify(
    request: Request,
    hub_mode: str = Query("", alias="hub.mode"),
    hub_challenge: str = Query("", alias="hub.challenge"),
    hub_verify_token: str = Query("", alias="hub.verify_token"),
):
    import os
    expected = os.getenv("WA_WEBHOOK_VERIFY_TOKEN", "shoptimize_wa_verify")
    if hub_mode == "subscribe" and hub_verify_token == expected:
        return Response(content=hub_challenge, media_type="text/plain")
    return JSONResponse({"ok": False}, status_code=403)


@router.post("/api/wa/webhook")
async def wa_webhook_incoming(request: Request):
    """Meta WA gelen mesaj webhook'u — opt-out işler."""
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": True})
    try:
        from services.wa_sender import handle_incoming_message
        for entry in body.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})
                phone_number_id = value.get("metadata", {}).get("phone_number_id", "")
                # Teşhis: her webhook'ta phone_id + kaç mesaj / kaç status geldi
                _msgs = value.get("messages", []) or []
                _stats = value.get("statuses", []) or []
                logger.info("[WA-HOOK] phone_id=%s mesaj=%d status=%d",
                            phone_number_id[-6:] if phone_number_id else "?", len(_msgs), len(_stats))
                for msg in value.get("messages", []):
                    from_phone = msg.get("from", "")
                    text_body  = (msg.get("text") or {}).get("body", "")
                    # Teşhis: gelen her mesajı logla (hangi numaraya, kimden, ne yazıldı)
                    logger.info("[WA-IN] phone_id=%s from=***%s text=%r",
                                phone_number_id[-6:] if phone_number_id else "?",
                                from_phone[-4:] if from_phone else "?", (text_body or "")[:40])
                    if from_phone and text_body:
                        await handle_incoming_message(phone_number_id, from_phone, text_body)
                # Teslim/okundu durumları → kampanya istatistiği (iletildi/okundu)
                for st in value.get("statuses", []):
                    msg_id = st.get("id", "")
                    status = st.get("status", "")  # sent | delivered | read | failed
                    if not msg_id or status not in ("delivered", "read"):
                        continue
                    owner = await store.resolve_campaign_message(msg_id)
                    if owner:
                        u, b, cid = owner
                        await store.mark_campaign_delivery(u, b, cid, msg_id, status)
    except Exception as e:
        logger.warning("[WA] Webhook işleme hatası: %s", e)
    return JSONResponse({"ok": True})
